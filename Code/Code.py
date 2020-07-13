import tkinter as tk
from tkinter import ttk
from tkinter import PhotoImage
from PIL import Image, ImageTk
from openpyxl import *
from datetime import datetime
from win10toast import ToastNotifier

#Activation of Excel Workbook for data storage
wb = load_workbook(r'F:\Semester II - Spring 2020\CS 102 - Data Structures & Algorithms\Project\project-team-ibrahim\Code\Database.xlsx')
sheet = wb.active

#Data Structure; List will be used to manage the whole app
space = []

#Adding the previous data in the structure
for i in range(2, sheet.max_row+1):
    temp = []
    for j in range(1, sheet.max_column+1):
        temp.append(sheet.cell(i, j).value)
    space.append(temp)

#This function will standardise the Excel sheet with the required headers.
def excel(sheet):

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 10
    sheet.column_dimensions['H'].width = 10
    sheet.column_dimensions['I'].width = 40

    sheet.cell(row=1, column=1).value = "Task Name"
    sheet.cell(row=1, column=2).value = "Urgency"
    sheet.cell(row=1, column=3).value = "Starting Date"
    sheet.cell(row=1, column=4).value = "Starting Time"
    sheet.cell(row=1, column=5).value = "Ending Date"
    sheet.cell(row=1, column=6).value = "Ending Time"
    sheet.cell(row=1, column=7).value = "Reminder"
    sheet.cell(row=1, column=8).value = "Status"
    sheet.cell(row=1, column=9).value = "Description"

#Standardise Excel Sheet only if no data is extracted
if not space:
    excel(sheet)

#Functions to set focus on the Add Task window

def focus1(event): 
    urgency_field.focus_set() 

def focus2(event): 
    findate_field.focus_set() 

def focus3(event): 
    fintime_field.focus_set() 

def focus4(event):
    remind_field.focus_set() 

def focus5(event): 
    status_field.focus_set() 

def focus6(event): 
    note_field.focus_set() 

#Function to clear the content of data entry boxes after submission of one data packet and bring cursor back on first entry box
def clear():   
    taskname_field.delete(0, tk.END) 
    urgency_field.delete(0, tk.END) 
    findate_field.delete(0, tk.END) 
    fintime_field.delete(0, tk.END) 
    remind_field.delete(0, tk.END) 
    status_field.delete(0, tk.END) 
    note_field.delete(0, tk.END)
    taskname_field.focus_set()
  
#Function to take data from Add Task Window and save it in the data structure
def insert():
        var1 = taskname_field.get()
        var2 = urgency_field.get()
        if not var2 in ['1', '0']:
            notify = tk.Label(tp1, text="Use only 0 or 1 to indicate urgency.              ", fg="black")
            notify.place(x=100, y=300)
            return()
        var3 = datetime.now()
        var3 = var3.strftime("%Y/%m/%d %H:%M:%S")
        var3 = var3.split()
        var5 = findate_field.get()
        formatd = "%Y/%m/%d"
        try:
            datetime.strptime(var5, formatd)
        except ValueError:
            notify = tk.Label(tp1, text="Date format should be YYYY/MM/DD.                 ", fg="black")
            notify.place(x=100, y=300)
            return()
        var6 = fintime_field.get()
        var6 = var6 + ":00"
        formatt = "%H:%M:%S"
        try:
            datetime.strptime(var6, formatt)
        except ValueError:
            notify = tk.Label(tp1, text="Time format should be HH:MM.                      ", fg="black")
            notify.place(x=100, y=300)
            return()
        var7 = remind_field.get()
        if not var7 in ['1', '0']:
            notify = tk.Label(tp1, text="Use only 0 or 1 to indicate reminder.             ", fg="black")
            notify.place(x=100, y=300)
            return()
        var8 = status_field.get()
        if var8 == "":
            notify = tk.Label(tp1, text="Empty Input                                       ", fg="black")
            notify.place(x=100, y=300)
            return()
        elif not int(var8) in list(range(100)):
            notify = tk.Label(tp1, text="Use numbers between 0-100 to indicate status.     ", fg="black")
            notify.place(x=100, y=300)
            return()
        var9 = note_field.get()
        temp = [var1, var2, var3[0], var3[1], var5, var6, var7, var8, var9]
        space.append(temp)
        clear()

#Function to open the Add Task Window labelled as Taskpane 01
def taskpane1():
        global taskname_field, urgency_field, findate_field, fintime_field, remind_field, status_field, note_field, tp1

        #Creating Window
        tp1 = tk.Toplevel(m)
        tp1.geometry("650x400")
        tp1["bg"] = "red4"

        #Placing Buttons
        close = tk.Button(tp1, text = 'Close', width=25, command = tp1.destroy)
        close.place(x=240,y=350)

        #Preparing the GUI for submitting tasks
        heading = tk.Label(tp1, text="Add your tasks here.", font=("Helvetica",15))
        heading.place(x=250, y=10)

        taskname = tk.Label(tp1, text="Name your task: ", fg="black") 
        urgency = tk.Label(tp1, text="Is it an urgent task ? 1: Yes 0: No ", fg="black")
        findate = tk.Label(tp1, text="Enter the finishing date: YYYY/MM/DD ", fg="black") 
        fintime = tk.Label(tp1, text="Enter the finishing time: HH:MM ", fg="black")
        remind = tk.Label(tp1, text="Do you want to get reminded by email ? 1: Yes 0: No ", fg="black")
        status = tk.Label(tp1, text="How much task is completed till now ?", fg="black")
        note = tk.Label(tp1, text="Enter any additional note you want to. ", fg="black") 

        taskname.place(x=20, y=70)
        urgency.place(x=20, y=90)
        findate.place(x=20, y=110)
        fintime.place(x=20, y=130)
        remind.place(x=20, y=150)
        status.place(x=20, y=170)
        note.place(x=20, y=190)

        taskname_field = tk.Entry(tp1, width=50) 
        urgency_field = tk.Entry(tp1, width=50) 
        findate_field = tk.Entry(tp1, width=50) 
        fintime_field = tk.Entry(tp1, width=50) 
        remind_field = tk.Entry(tp1, width=50) 
        status_field = tk.Entry(tp1, width=50) 
        note_field = tk.Entry(tp1, width=50) 
  
        taskname_field.bind("<Return>", focus1)  
        urgency_field.bind("<Return>", focus2) 
        findate_field.bind("<Return>", focus3) 
        fintime_field.bind("<Return>", focus4) 
        remind_field.bind("<Return>", focus5) 
        status_field.bind("<Return>", focus6) 

        taskname_field.place(x=320, y=70)
        urgency_field.place(x=320, y=90)
        findate_field.place(x=320, y=110)
        fintime_field.place(x=320, y=130)
        remind_field.place(x=320, y=150)
        status_field.place(x=320, y=170)
        note_field.place(x=320, y=190)

        submit = tk.Button(tp1, text="Submit", fg="Black", bg="Red", command=insert)
        submit.place(x=450, y=220)

#Function to open the Remove Task Window labelled as Taskpane 02
def taskpane2():
        #Creating Window
        global listBox2, tp2
        tp2 = tk.Toplevel(m)
        tp2.geometry("700x400")
        tp2["bg"] = "red4"

        #Creating Buttons
        refresh = tk.Button(tp2, text = 'Refresh', width=25, command = refresh2)
        refresh.place(x=100,y=350)

        close = tk.Button(tp2, text = 'Close', width=25, command = tp2.destroy)
        close.place(x=420,y=350)

        #Preparing the GUI for removing tasks
        heading = tk.Label(tp2, text="Remove your tasks here.", font=("Helvetica",15))
        heading.place(x=240, y=10)

        cols = ('Task', 'Starting Date', 'Starting Time', 'Deadline', 'Target Hour', 'Status')

        listBox2 = ttk.Treeview(tp2, columns=cols, show='headings')
        listBox2.place(x=50, y=50)

        for col in cols:
            x = 100
            listBox2.heading(col, text=col)
            listBox2.column(col, minwidth=0, width=x)
            x += 100
        for item in space:
            listBox2.insert("", "end", values=(item[0], item[2], item[3], item[4], item[5], item[7]))
        listBox2.bind('<Double-1>', remove)

#Function to open the Sort Task Window labelled as Taskpane 03
def taskpane3():
        #Creating Window
        global listBox3, tp3
        tp3 = tk.Toplevel(m)
        tp3.geometry("700x480")
        tp3["bg"] = "red4"

        #Creating Buttons
        refresh = tk.Button(tp3, text = 'Refresh', width=25, command = refresh3)
        refresh.place(x=100,y=430)

        close = tk.Button(tp3, text = 'Close', width=25, command = tp3.destroy)
        close.place(x=420,y=430)

        #Preparing the GUI for sorting tasks
        heading = tk.Label(tp3, text="Sort out your tasks here.", font=("Helvetica",15))
        heading.place(x=240, y=10)
        
        sort1 = tk.Button(tp3, text = 'Sort through name.', width=25, command = lambda: quicksort(space, 0, len(space)-1, 0))
        sort1.place(x=160,y=300)
        sort2 = tk.Button(tp3, text = 'Sort through start date', width=25, command = lambda: quicksort(space, 0, len(space)-1, 2))
        sort2.place(x=160,y=330)
        sort3 = tk.Button(tp3, text = 'Sort through deadline', width=25, command = lambda: quicksort(space, 0, len(space)-1, 4))
        sort3.place(x=350,y=300)
        sort4 = tk.Button(tp3, text = 'Sort through target hour', width=25, command = lambda: quicksort(space, 0, len(space)-1, 5))
        sort4.place(x=350,y=330)
        sort5 = tk.Button(tp3, text = 'Show urgent tasks', width=25, command = lambda: quicksort(space, 0, len(space)-1, 1))
        sort5.place(x=250,y=360)

        cols = ('Task', 'Starting Date', 'Starting Time', 'Deadline', 'Target Hour', 'Status')

        listBox3 = ttk.Treeview(tp3, columns=cols, show='headings')
        listBox3.place(x=50, y=50)
        vsb = ttk.Scrollbar(tp3, orient="vertical", command=listBox3.yview)
        vsb.place(x=650+2, y=50, height=200+26)

        for col in cols:
            x = 100
            listBox3.heading(col, text=col)
            listBox3.column(col, minwidth=0, width=x)
            x += 100
        for item in space:
            listBox3.insert("", "end", values=(item[0], item[2], item[3], item[4], item[5], item[7]))

#Function to present information about the clicked task - Main Window
def info(event):
    item = listBox.focus()
    label = tk.Label(m, text="Task Details", font=("Helvetica",15), fg='red4').place(x=700,y=150)
    d1 = tk.Label(m, text=str(listBox.item(item)['values'][0]), font=("Arial",12)).place(x=700,y=190)
    d2 = tk.Label(m, text=str(listBox.item(item)['values'][6]), font=("Arial",8)).place(x=700,y=220)

#Function to refresh the tasks information on Main Window
def refresh1():
    listBox.delete(*listBox.get_children())
    for item in space:
        listBox.insert("", "end", values=(item[0], item[2], item[3], item[4], item[5], item[7], item[8]))
    ender()

#Function to refresh the tasks information on Remove Task Window
def refresh2():
    listBox2.delete(*listBox2.get_children())
    for item in space:
        listBox2.insert("", "end", values=(item[0], item[2], item[3], item[4], item[5], item[7]))
    ender()

#Function to refresh the tasks information on Sort Task Window
def refresh3():
    listBox3.delete(*listBox3.get_children())
    for item in space:
        listBox3.insert("", "end", values=(item[0], item[2], item[3], item[4], item[5], item[7]))
    ender()

#Functions to search the task to remove - Use of Binary Search
def binary(lst, left, right, x):
  mid = (left + right)//2
  temp = lst[mid][3]
  temp = temp.split(':')
  temp = ''.join(temp)
  temp = int(temp)
  if temp == x:
    return mid
  else:
    if temp > x:
      return binary(lst, left, mid-1, x)
    elif temp < x:
      return binary(lst, mid+1, right, x)

def remove(event):
    item = listBox2.focus()
    x = listBox2.item(item)['values'][2]
    x = x.split(':')
    x = ''.join(x)
    x = int(x)
    quicksort(space, 0, len(space)-1, 3)
    if len(space) == 1:
        del(space[0])
    else:
        del(space[binary(space, 0, len(space), x)])
    removal = tk.Label(tp2, text="Item removed !").place(x=320, y=300)

#Functions to sort the list of tasks - Use of Quick Sort Algorithm
def partition0(lst, x, y, typeindex):
  i = x - 1
  for j in range(x, y):
    if lst[j][typeindex] <= lst[y][typeindex]:
      i += 1
      lst[i], lst[j] = lst[j], lst[i]
  lst[i+1], lst[y] = lst[y], lst[i+1]
  return i + 1

def partition1(lst, x, y, typeindex):
  i = x - 1
  for j in range(x, y):
    if numery1(lst[j][typeindex]) <= numery1(lst[y][typeindex]):
      i += 1
      lst[i], lst[j] = lst[j], lst[i]
  lst[i+1], lst[y] = lst[y], lst[i+1]
  return i + 1

def partition2(lst, x, y, typeindex):
  i = x - 1
  for j in range(x, y):
    if lst[j][typeindex] >= lst[y][typeindex]:
      i += 1
      lst[i], lst[j] = lst[j], lst[i]
  lst[i+1], lst[y] = lst[y], lst[i+1]
  return i + 1

def partition3(lst, x, y, typeindex):
  i = x - 1
  for j in range(x, y):
    if numery2(lst[j][typeindex]) <= numery2(lst[y][typeindex]):
      i += 1
      lst[i], lst[j] = lst[j], lst[i]
  lst[i+1], lst[y] = lst[y], lst[i+1]
  return i + 1

def quicksort(lst, x, y, typeindex):
  if not x < y:
    return
  if typeindex == 0:
    z = partition0(lst, x, y, typeindex)
  elif typeindex in [1, 6]:
    z = partition2(lst, x, y, typeindex)
  elif typeindex in [3, 5]:
    z = partition3(lst, x, y, typeindex)
  else:
    z = partition1(lst, x, y, typeindex)
  quicksort(lst, x, z - 1, typeindex)
  quicksort(lst, z + 1, y, typeindex)

#Giving out reminders

quicksort(space, 0, len(space)-1, 6)
x = 0
string = 'Do not forget: '
while space[x][6] == '1':
    string += space[x][0]
    string += ', '
    x += 1
toaster = ToastNotifier()
toaster.show_toast("PocketList Task Reminder", string[:-1], duration=10, threaded=True)

#Extra Functions: Numery1 and Numery2 turns date and time into integers respectively
def numery1(x):
    x = x.split('/')
    x = ''.join(x)
    x = int(x)
    return x

def numery2(x):
    x = x.split(':')
    x = ''.join(x)
    x = int(x)
    return x

#Function to save data in the database
def ender():
    del wb['Sheet1']
    sheet = wb.create_sheet('Sheet1')
    excel(sheet)
    for j in range(0, len(space)):
            sheet.cell(row=j+2, column=1).value = space[j][0]
            sheet.cell(row=j+2, column=2).value = space[j][1] 
            sheet.cell(row=j+2, column=3).value = space[j][2]
            sheet.cell(row=j+2, column=4).value = space[j][3] 
            sheet.cell(row=j+2, column=5).value = space[j][4] 
            sheet.cell(row=j+2, column=6).value = space[j][5]
            sheet.cell(row=j+2, column=7).value = space[j][6]
            sheet.cell(row=j+2, column=8).value = space[j][7] 
            sheet.cell(row=j+2, column=9).value = space[j][8]
    wb.save(r'F:\Semester II - Spring 2020\CS 102 - Data Structures & Algorithms\Project\project-team-ibrahim\Code\Database.xlsx')

#The Main Window

#Setting up the window
m = tk.Tk(screenName=None,  baseName=None,  className='PocketList',  useTk=1)
m.title('PocketList')
m.geometry('1000x400')

#Setting up the buttons
AddTask = tk.Button(m, text = 'Add Task', width=25, command = taskpane1).place(x=700,y=50)
RemoveTask = tk.Button(m, text = 'Remove Task', width=25, command = taskpane2).place(x=700,y=80)
SortTask = tk.Button(m, text = 'Sort Tasks', width=25, command = taskpane3).place(x=700,y=110)
Refresh = tk.Button(m, text="Refresh", width=20, command=refresh1).place(x=150,y=300)
CloseButton = tk.Button(m, text="Close", width=20, command=m.destroy).place(x=350,y=300)

path = r"F:\Semester II - Spring 2020\CS 102 - Data Structures & Algorithms\Project\project-team-ibrahim\Resources\Brand Mark\PocketList - Footer.png"

#Creates a Tkinter-compatible photo image, which can be used everywhere Tkinter expects an image object.
img = ImageTk.PhotoImage(Image.open(path))

#The Label widget is a standard Tkinter widget used to display a text or image on the screen.
panel = tk.Label(m, image = img)

#The Pack geometry manager packs widgets in rows or columns.
panel.place(x=0, y=350)

#Preparing the GUI for this window
Dashboard = tk.Label(m, text="Dashboard", font=("Helvetica",15)).place(x=300,y=10)
Options = tk.Label(m, text="Options", font=("Helvetica",15)).place(x=750,y=10)

cols = ('Task', 'Starting Date', 'Starting Time', 'Deadline', 'Target Hour', 'Status')

listBox = ttk.Treeview(m, columns=cols, show='headings')
listBox.place(x=50, y=50)
vsb = ttk.Scrollbar(m, orient="vertical", command=listBox.yview)
vsb.place(x=650+2, y=50, height=200+26)

for col in cols:
    x = 100
    listBox.heading(col, text=col)
    listBox.column(col, minwidth=0, width=x)
    x += 100
for item in space:
    listBox.insert("", "end", values=(item[0], item[2], item[3], item[4], item[5], item[7], item[8]))

listBox.bind('<Double-1>', info)

m.mainloop()
